import openpyxl
from openpyxl import Workbook

class Kielletyt():
    def __init__(self, list):
        self.list = list
        self.list = self.avaaKielletytKategoriat()
        print(list)
    def avaaKielletytKategoriat(self):
        self.list = open('C:/pythontestit/' + 'kielletyt_kategoriat_duunitori.txt', encoding='utf-8').read().split('\n')
        print(self.list)
        print("kieltolista luotu")
    def KiellettyjenLisays(self):
        print("looppiin")
        #avataan excelin uudet arvot
        wb = openpyxl.load_workbook('C:\pythontestit\Hakutulokset.xlsx', read_only=True) #täältä otetaan kielletyt ja lisätään listalle
        ws = wb.active
        #seuraavaksi tarkistetaan onko excelissä nollia
        for row in ws.iter_rows(min_col=7, max_col=7):
            for cell in row:
                if cell.value == 0:
                    #jos on nollia, onko kategoria jo kiellettyjen joukossa
                    #print(ws.cell(row=cell.row, column=6).value)
                    #print(self.list.split('\n'))
                    if ws.cell(row=cell.row, column=6).value not in self.list:
                        self.list.append(str(ws.cell(row=cell.row, column=6).value))
                        print(list)
        apu = ""
        for item in list:
            apu += item + '\n'
        kakka = open('C:/pythontestit/' + 'kielletyt_kategoriat_duunitori.txt', 'w+', encoding='utf-8')
        kakka.write(apu)
        kakka.close()
kieltolista = []
uusiKieltolista = []
kieltolista = Kielletyt(kieltolista)
print(kieltolista)
#uusiKieltolista = Kielletyt.KiellettyjenLisays(uusiKieltolista)