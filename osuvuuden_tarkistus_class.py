from openpyxl import Workbook
import openpyxl
#pitäskö duunitorin kategorioita käsittelevät työntää samaan luokkaan ja kielletyt hakusanat toiseen luokkaan?
#kielletyt hakusanat tarkistettaisiin myös työkkärin haussa

class Kielletyt():
    def __init__(self):
        self.luoLista()
    def luoLista(self):
        self.list = open('C:/pythontestit/' + 'kielletyt_kategoriat_duunitori.txt', encoding='utf-8').read().split('\n')
    def kategorian_tarkistus(self, kategoria):
        if kategoria in self.list: #toimii
            #print(kategoria)
            return True
        else:
            return False

def osuvuudet():
    kieltolista = open('C:/pythontestit/' + 'kielletyt_kategoriat_duunitori.txt', encoding='utf-8').read().split('\n')
    print(1)
    print(kieltolista)
    wb = openpyxl.load_workbook('C:\pythontestit\Hakutulokset.xlsx', read_only=True) #täältä otetaan kielletyt ja lisätään listalle
    ws = wb.active
    for row in ws.iter_rows(min_col=7, max_col=7):
        for cell in row:
            if cell.value == 0:
                #jos on nollia, onko kategoria jo kiellettyjen joukossa
                #print(ws.cell(row=cell.row, column=6).value)
                if ws.cell(row=cell.row, column=6).value not in kieltolista:
                    kieltolista.append(str(ws.cell(row=cell.row, column=6).value))
    apu = ""
    for item in kieltolista:
        apu += item + '\n' #saisko tän forin kanssa samaan riviin?
    print(3)
    print(kieltolista)
    istunto = open('C:/pythontestit/' + 'kielletyt_kategoriat_duunitori.txt', 'w+', encoding='utf-8')
    print(istunto)
    istunto.write(apu)
    istunto.close()
    print("valmis")
    return;