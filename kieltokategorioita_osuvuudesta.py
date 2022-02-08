import openpyxl
from openpyxl import Workbook

input = "erikoisalat (terveydenhuolto)"
#ota seuraavaksi jo kielletyt kategoriat
kieltolista = open('C:/pythontestit/' + 'kielletyt_kategoriat_duunitori.txt', encoding='utf-8').read().split('\n')
if input in kieltolista:
    print("on kieltolistalla")
#avataan excelin uudet arvot
#wb = openpyxl.load_workbook('C:\pythontestit\Hakutulokset.xlsx', read_only=True) #täältä otetaan kielletyt ja lisätään listalle
#ws = wb.active
#seuraavaksi tarkistetaan onko excelissä nollia
#for row in ws.iter_rows(min_col=7, max_col=7):
    #for cell in row:
        #if cell.value == 0:
            #jos on nollia, onko kategoria jo kiellettyjen joukossa
            #print(ws.cell(row=cell.row, column=6).value)
            #if ws.cell(row=cell.row, column=6).value not in kieltolista:
                #kieltolista.append(str(ws.cell(row=cell.row, column=6).value))
                #print(kieltolista)
#apu = ""
for item in kieltolista:
    #apu += item + '\n'
    if input == item:
        print("löytyi")
#kakka = open('C:/pythontestit/' + 'kielletyt_kategoriat_duunitori.txt', 'w+', encoding='utf-8')
#kakka.write(apu)
#kakka.close()
