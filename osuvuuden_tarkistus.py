from openpyxl import Workbook
import openpyxl
import access_files
#pitäskö duunitorin kategorioita käsittelevät työntää samaan luokkaan ja kielletyt hakusanat toiseen luokkaan?
#kielletyt hakusanat tarkistettaisiin myös työkkärin haussa

polku = 'duunitorin_kielletyt_kategoriat.json'

#avaa kieltolista lukumuodossa
def kieltolista_luku():
     return access_files.open_json(polku)

def osuvuudet():
    kielletyt = kieltolista_luku()
    wb = openpyxl.load_workbook('Hakutulokset.xlsx', read_only=True) # otetaan kielletyt ja listalle
    ws = wb.active
    for row in ws.iter_rows(min_col=7, max_col=7):
        for cell in row:
            if cell.value == 0:
                #jos on nollia, onko kategoria jo kiellettyjen joukossa
                if ws.cell(row=cell.row, column=6).value not in kielletyt:
                    kielletyt.append(str(ws.cell(row=cell.row, column=6).value))
    access_files.write_data_to_json_with(kielletyt, polku)
    return